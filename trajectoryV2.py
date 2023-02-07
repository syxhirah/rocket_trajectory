def rotation_matric(Q):
    """ Obtain rotation matrix from corresponding quaternion """
    s, vx, vy, vz = Q
    return np.array([[1 - 2 * vy ** 2 - 2 * vz ** 2, 2 * vx * vy - 2 * s * vz, 2 * vx * vz + 2 * s * vy],
                     [2 * vx * vy + 2 * s * vz, 1 - 2 * vx ** 2 - 2 * vz ** 2, 2 * vy * vz - 2 * s * vx],
                     [2 * vx * vz - 2 * s * vy, 2 * vy * vz + 2 * s * vx, 1 - 2 * vx ** 2 - 2 * vy ** 2]])

def normalize(v):
    """ Normalize vector """
    norm = np.linalg.norm(v)
    if norm == 0:
        return v
    return v/norm


class Trajectory:
    def __init__(self):
        self.dt = 0.1
        # Rocket Parameters
        self.m_empty = 1.950                                        # empty mass of the rocket [kg]
        self.radius = 0.07                                          # rocket radius [m]
        self.X_RB = 1.07                                            # rocket length [m]
        self.A_RB = 0.015                                           # cross-sectional area [m^2]
        self.X_cm0 = 0.677                                          # initial center of mass (distance from nose tip) [m]
        self.X_cmf = 0.272                                          # final center of mass (distance from nose tip) [m]
        self.X_cp = 0.499                                           # center of pressure (distance from nose tip) [m]
        self.C_D = 0.2                                              # drag coefficient [-]

        # Rocket Reference Coordinate System
        self.Y_A0 = np.array([1, 0, 0])                                         # reference yaw axis
        self.P_A0 = np.array([0, 1, 0])                                         # reference pitch axis
        self.R_A0 = np.array([0, 0, 1])                                         # reference roll axis
        self.launchAngle = 30                                                  # launch angle from vertical [deg]
        self.R_init = np.array([[1, 0, 0],                                      # initial orientation
                                [0, np.cos(radians(self.launchAngle)), np.sin(radians(self.launchAngle))],
                                [0, -np.sin(radians(self.launchAngle)), np.cos(radians(self.launchAngle))]])

        # Load motor data
        rse_data = requests.get('https://drive.google.com/uc?export=download&id=1EJWwxQ53LCQqO5ysx2e5KISwcTwEB0mP').text
        #https://drive.google.com/file/d/1EJWwxQ53LCQqO5ysx2e5KISwcTwEB0mP/view?usp=sharing
        engine = thrustcurve.loads(rse_data)
        e = engine[0]

        t_raw = e.data['time']                          # time-values for engine data
        m_raw = e.data['mass']                          # engine mass; measured in grams at t_raw
        T_raw = e.data['force']                         # thrust; measured in newtons at t_raw

        t_raw = t_raw.to_numpy()
        m_raw = m_raw.to_numpy() / 1000 + self.m_empty  # convert m_raw to kg and add dry weight of the rocket
        T_raw = T_raw.to_numpy()

        self.t_burn = t_raw[-1]
        self.f_mass = interpol.interp1d(t_raw, m_raw, kind='linear', fill_value=(m_raw[0], m_raw[-1]), bounds_error=False)
        self.f_thrust = interpol.interp1d(t_raw, T_raw, kind='linear', fill_value=(0, 0), bounds_error=False)

        # Rocket state at time t
        self.t = 0
        self.X = np.array([0, 0, 0])
        self.Q = np.array([1, 0, 0, 0])
        self.P = np.array([0, 0, 0])
        self.L = np.array([0, 0, 0])

        # Simulation state
        self.lift_off = False
        self.ascent = False
        self.apogee = False
        self.descent = False
        self.touch_down = False

        # Data arrays
        self.x_vec = []
        self.t_vec = []
        self.V_vec = []
        
        self.coord = []
        self.x_coor = []
        self.y_coor = []
        self.vx = []
        self.vy = []
        self.V = []
        self.long = []
        self.lat = []
        self.wgs = []

        # Launch coordinates: 6.317052229732824, 100.4736655882026
        self.inilat = 6.317052229732824      #latitude x
        self.inilong = 100.4736655882026   #longitude y

    def simulation(self):
        """ Function which carries out the simulation loop, solving the system of ODEs through time """      
        while not self.touch_down:
            self.x_vec.append(self.X)
            self.t_vec.append(self.t)
            self.RKF45()
            
            # At apogee
            if self.ascent and len(self.x_vec) > 1:
                if self.x_vec[-1][2] < self.x_vec[-2][2]:
                    self.apogee = True
                    self.descent = True
                    self.ascent = False
            # At touch-down
            if self.descent:
                if self.X[2] < 0:
                    self.touch_down = True          
           
    def RKF45(self):
        """ function to carry out role of numerically solving system of ODEs using the Runge-Kutta 45 method """
        y = np.array([self.X, self.Q, self.P, self.L], dtype=object)

        # calculate k values for all ODE
        k1 = self.dt * self.rocket_dynamics(self.t, y)
        k2 = self.dt * self.rocket_dynamics(self.t + 1 / 2 * self.dt, y + 1 / 2 * k1)
        k3 = self.dt * self.rocket_dynamics(self.t + 1 / 2 * self.dt, y + 1 / 2 * k2)
        k4 = self.dt * self.rocket_dynamics(self.t + self.dt, y + k3)

        # calculate next state
        y_next = y + 1 / 6 * k1 + 1 / 3 * k2 + 1 / 3 * k3 + 1 / 6 * k4

        # update time and current state
        self.t = self.t + self.dt
        self.X, self.Q, self.P, self.L = y_next

    def rocket_dynamics(self, t, y):
        """ function to carry out role of the Rocket Dynamic Model; calculate state derivatives -- RHS of system of ODEs """
        derivatives = []
        X, Q, P, L = y
        x, y, z = X
        # standard parameters
        rho = Atmosphere(float(z)).density
        g = 9.81
        m = self.f_mass(t)
        X_cm = self.centre_mass(t)
        X_bar = (self.X_cp - X_cm)

        # Rocket orientation
        Q = Q / np.linalg.norm(Q)
        s = Q[0]
        v = Q[1:]
        R = rotation_matric(Q)
        R_coord = R.dot(self.R_init)
        YA = R_coord.dot(self.Y_A0) / np.linalg.norm(R_coord.dot(self.Y_A0))
        PA = R_coord.dot(self.P_A0) / np.linalg.norm(R_coord.dot(self.P_A0))
        RA = R_coord.dot(self.R_A0) / np.linalg.norm(R_coord.dot(self.R_A0))

        # Position derivative
        X_dot = P / m
        derivatives.append(np.array(X_dot))

        # Quaternion derivative
        I0 = self.MMOI(t)
        omega = R.dot(np.linalg.inv(I0)).dot(np.transpose(R)).dot(np.transpose(L))
        omega_hat = normalize(omega)
        s_dot = -1 / 2 * np.dot(omega, v)
        v_dot = 1 / 2 * (s * omega + (np.cross(omega, v)))

        derivatives.append(np.array(np.append(s_dot, v_dot)))

        # Calculate Velocity
        V_cm = X_dot
        V_omega = X_bar * np.sin(np.arccos(np.dot(RA, omega_hat))) * (np.cross(RA, omega_hat))

        V = V_cm + V_omega
        V_hat = normalize(V)
        V_mag = np.linalg.norm(V)

        # Force F(t)
        F_thrust = self.f_thrust(t) * RA
        F_gravity = np.array([0, 0, -m * g])
        F_drag = -1 / 2 * rho * V_mag * self.A_RB * self.C_D * V
        F_drag_axial = np.dot(F_drag, RA) * RA
        F_drag_normal = F_drag - F_drag_axial

        if abs(F_thrust[2]) > abs(F_gravity[2]):
            self.lift_off = True
            self.ascent = True
        F = F_drag_axial + F_drag_normal + F_gravity + F_thrust if self.lift_off else [0, 0, 0]

        derivatives.append(np.array(F))

        # Moments M(t)
        moment = np.linalg.norm(F_drag_normal) * X_bar * np.cross(RA, V_hat)

        derivatives.append(np.array(moment))

        # Collect velocity data
        self.V_vec.append(V)

        return np.array(derivatives, dtype=object)

    def centre_mass(self, t):
        X_cm = (self.X_cmf - self.X_cm0)/(self.m_empty - self.f_mass(0))*(self.f_mass(t) - self.f_mass(0)) + self.X_cm0
        return X_cm

    def MMOI(self, t):
        # mass moment of inertia based on current mass and with the rocket approximated as a cylinder
        m =self.f_mass(t)
        Ixx = 1/4*m*self.radius**2+1/12*m*self.X_RB**2
        Iyy = 1/4*m*self.radius**2+1/12*m*self.X_RB**2
        Izz = 1/2*m*self.radius**2
        I0 = np.array([[Ixx, 0, 0],
                       [0, Iyy, 0],
                       [0, 0, Izz]])
        return I0
        
    def plot(self):
        
        pos = np.array(self.x_vec)

        f = plt.figure()
        f.set_figwidth(8)
        f.set_figheight(4)
        plt.plot(pos[:, 1], pos[:, 2])
        plt.xlabel('Horizontal Displacement [m]')
        plt.ylabel('Altitude [m]')
        plt.grid()
        plt.savefig("xvy.svg", format="svg")
        plt.show()

        f = plt.figure()
        f.set_figwidth(8)
        f.set_figheight(4)
        plt.plot(self.t_vec, pos[:, 2])
        plt.xlabel('Time [s]')
        plt.ylabel('Altitude [m]')
        plt.grid()
        plt.savefig("tvy.svg", format="svg")
        plt.show()

        f = plt.figure()
        f.set_figwidth(8)
        f.set_figheight(4)
        plt.plot(self.t_vec, self.vx)
        plt.xlabel('Time [s]')
        plt.ylabel('Horizontal Velocity [m/s]')
        plt.grid()
        plt.savefig("tvvx.svg", format="svg")
        plt.show()

        f = plt.figure()
        f.set_figwidth(8)
        f.set_figheight(4)
        plt.plot(self.t_vec, self.vy)
        plt.xlabel('Time [s]')
        plt.ylabel('Vertical Velocity [m/s]')
        plt.grid()
        plt.savefig("tvvy.svg", format="svg")
        plt.show()

        f = plt.figure()
        f.set_figwidth(8)
        f.set_figheight(4)
        plt.plot(self.t_vec, self.V)
        plt.xlabel('Time [s]')
        plt.ylabel('Velocity [m/s]')
        plt.grid()
        plt.savefig("tvv.svg", format="svg")
        plt.show()

        
    def Run(self):
        self.simulation()

    def data(self):
        Vtemp = []
        vxtemp = []
        vytemp = []
        for i in range(len(self.x_vec)):
            self.x_coor.append(self.x_vec[i][1])
            self.y_coor.append(self.x_vec[i][2])
            i = i+1
        for i in range(len(self.V_vec)):
            vx = self.V_vec[i][1]
            vy = self.V_vec[i][2]
            actualV = np.sqrt((vx**2)+(vy**2))
            Vtemp.append(actualV)
            vxtemp.append(vx)
            vytemp.append(vy)
            i = i+1
        for i in range(0, len(Vtemp), 4):
            self.V.append(Vtemp[i])
            self.vx.append(vxtemp[i])
            self.vy.append(vytemp[i])

    def coordinates(self):
        for i in range(len(self.x_coor)):
            new_lat = self.inilat 
            new_long= self.inilong+ (((self.x_coor[i])/6378000)*((180/np.pi))/np.cos(self.inilat*(np.pi/180)))
            arr = (new_long, new_lat, self.x_coor[i])
            self.lat.append(new_lat)
            self.long.append(new_long)
            self.wgs.append(arr)
            i = i+1
            
    def output(self):
        print("\nLaunch angle (deg):", self.launchAngle)
        print("Range (m):", self.x_coor[-1])
        print("Apogee (m):", max(self.y_coor))
        print("Maximum velocity (m/s):", max(self.V))
        print("Total time of flight (s):", self.t_vec[-1])
        print("\nLaunch coordinate:", self.inilat, ",", self.inilong)
        print("Touchdown coordinate:", self.lat[-1], ",", self.long[-1], "\n")
        
    def xlsx(self):    
        wb = Workbook()
        ws = wb.active

        for i in range(0,len(self.x_coor)):
            ws.cell(row=i+1,column=1).value=self.t_vec[i]   #time
            ws.cell(row=i+1,column=2).value=self.x_coor[i]  #x coordinate
            ws.cell(row=i+1,column=3).value=self.y_coor[i]  #y coordinate
            ws.cell(row=i+1,column=4).value=self.V[i]       #velocity
            ws.cell(row=i+1,column=5).value=self.vx[i]      #x velocity
            ws.cell(row=i+1,column=6).value=self.vy[i]      #y velocity
            ws.cell(row=i+1,column=7).value=self.lat[i]     #latitude
            ws.cell(row=i+1,column=8).value=self.long[i]    #longitude
        wb.save("UPM_Baseline.xlsx")

    def kml(self):
        f = open('UPM_Baseline.kml', 'w')
        f.write("""<?xml version="1.0" encoding="UTF-8"?>
<kml xmlns="http://earth.google.com/kml/2.2">
  <Document>
    <Style id="line">
      <LineStyle>
        <color>44ff00ff</color>
        <width>10</width>
      </LineStyle>
      <PolyStyle>
        <color>44ff00ff</color>
      </PolyStyle>
    </Style>
    <Placemark>
      <name>Flight Path AirAsia</name>
      <styleUrl>#line</styleUrl>
      <LineString>
        <extrude>1</extrude>
        <altitudeMode>absolute</altitudeMode>
        <coordinates>""")
        for i in range(len(self.lat)):
            f.write("        "+str(self.long[i]) + ","+ str(self.lat[i]) + "," + str(self.y_coor[i]) +"\n")
        f.write("""        </coordinates>
       </LineString>   
    </Placemark>
  </Document>
</kml>""")
        f.close()
        
if __name__ == '__main__':
    import numpy as np
    import matplotlib.pyplot as plt
    from ambiance import Atmosphere
    import thrustcurve
    import requests
    from scipy import interpolate as interpol
    from math import radians
    from openpyxl import Workbook

    trajectory = Trajectory()
    trajectory.Run()
    trajectory.data()
    trajectory.coordinates()
    trajectory.output()
    trajectory.plot()
    trajectory.xlsx()
    trajectory.kml()
